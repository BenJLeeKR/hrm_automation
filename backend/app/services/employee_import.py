"""사원 목록 Excel Import 서비스 (SCR-003 "인력마스터_ResourceTable" 시트 형식).

정책 확정 내용(로드맵 §7/§9 참조, 사용자 확정):
1. 팀/직급/역할/기술 명칭이 마스터 데이터에 없으면 자동 생성하지 않고, 해당 행만
   건너뛰지도 않는다 — 전체 Import를 실패 처리하고 행 번호/컬럼명/입력값/사유를 반환한다
   (오타·비표준 명칭이 마스터에 자동 유입되는 것을 방지).
2. `EMPL_NO`를 업무 기준키로 Upsert한다 — 기존 사번이면 수정, 없으면 신규 등록.
   단, 업로드 파일 내부에 동일 `EMPL_NO`가 중복되면 검증 오류로 전체 실패 처리한다.
3. 검증 단계에서 오류가 1건이라도 있으면 DB 변경 없이 실패 응답만 반환한다(전체 성공/
   전체 실패). 모든 행이 검증을 통과한 경우에만 하나의 트랜잭션으로 Upsert를 수행한다.
"""

import io
import uuid
from dataclasses import dataclass, field
from datetime import date, datetime

from openpyxl import load_workbook
from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from app.models.hr_dept_mst import HrDeptMst
from app.models.hr_empl_mst import HrEmplMst
from app.models.hr_empl_role_rel import HrEmplRoleRel
from app.models.hr_empl_skill_rel import HrEmplSkillRel
from app.models.hr_jikgup_mst import HrJikgupMst
from app.models.hr_jikmu_mst import HrJikmuMst
from app.models.hr_skill_mst import HrSkillMst

# Export API(`GET /api/v1/employees/export`)와 동일한 헤더 순서 — 왕복 호환을 위해 그대로 재사용
_HEADERS = ["사번", "성명", "팀", "직급", "보유역할", "주요기술", "숙련도", "입사일", "재직상태", "휴대폰번호"]
_EMPL_STAT_LABEL_TO_CD = {"재직": "ACTIVE", "휴직": "LEAVE", "퇴직": "RETIRED"}


class EmployeeImportValidationError(Exception):
    """검증 단계 실패 — DB에는 아무것도 반영하지 않는다."""

    def __init__(self, total_rows: int, errors: list[dict]):
        self.total_rows = total_rows
        self.errors = errors
        super().__init__(f"{len(errors)}건의 검증 오류")


@dataclass
class _ParsedRow:
    row_no: int
    empl_no: str
    empl_nm: str
    dept_id: uuid.UUID
    jikgup_id: uuid.UUID
    jikmu_ids: list[uuid.UUID] = field(default_factory=list)
    skill_ids: list[uuid.UUID] = field(default_factory=list)
    prfcy_levl: int | None = None
    hire_dt: date | None = None
    empl_stat_cd: str = "ACTIVE"
    mphone_no: str | None = None


def _err(row: int, column: str, value: object, reason: str) -> dict:
    return {"row": row, "column": column, "value": None if value is None else str(value), "reason": reason}


def _cell(row: tuple, index: int) -> str | None:
    value = row[index] if index < len(row) else None
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def parse_and_validate(db: Session, file_bytes: bytes) -> list[_ParsedRow]:
    """Excel 파일을 파싱하고 전체 행을 검증한다. 오류가 1건이라도 있으면
    `EmployeeImportValidationError`를 발생시키며, 이 시점까지는 DB에 아무 것도 쓰지 않는다.
    """
    try:
        workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as exc:  # noqa: BLE001 — 손상된 파일 등 openpyxl이 던지는 다양한 예외를 통일 처리
        raise EmployeeImportValidationError(0, [_err(1, "file", None, f"Excel 파일을 읽을 수 없습니다: {exc}")]) from exc

    sheet = workbook.active
    all_rows = list(sheet.iter_rows(values_only=True))
    if not all_rows:
        raise EmployeeImportValidationError(0, [_err(1, "file", None, "빈 파일입니다.")])

    header = [str(v).strip() if v is not None else "" for v in all_rows[0][: len(_HEADERS)]]
    if header != _HEADERS:
        _fail_header(all_rows, header)

    data_rows = all_rows[1:]
    errors: list[dict] = []

    # 마스터 데이터를 명칭→ID로 미리 전부 조회 — 행마다 개별 쿼리하지 않고 한 번에 매핑
    dept_by_name = {d.DEPT_NM: d.DEPT_ID for d in db.scalars(select(HrDeptMst))}
    jikgup_by_name = {j.JIKGUP_NM: j.JIKGUP_ID for j in db.scalars(select(HrJikgupMst))}
    jikmu_by_code = {j.JIKMU_CD: j.JIKMU_ID for j in db.scalars(select(HrJikmuMst))}
    skill_by_name = {s.SKILL_NM: s.SKILL_ID for s in db.scalars(select(HrSkillMst))}
    # `HR_EMPL_MST.EMAIL_ADDR`가 NOT NULL로 전환되었으나(2026-07-06 설계 확정, §8 큐 1-1)
    # 이 Import 양식(`_HEADERS`)에는 이메일 컬럼이 없다 — 기존 사원 수정(EMAIL_ADDR 미변경)은
    # 그대로 지원하되, 신규 등록은 이메일을 채울 방법이 없어 검증 오류로 막는다. 이메일 컬럼
    # 추가는 Excel 양식 자체를 바꾸는 더 큰 범위라 별도 후속 작업으로 분리(§9 리스크 참조).
    existing_empl_nos = {row[0] for row in db.execute(select(HrEmplMst.EMPL_NO)).all()}

    parsed_rows: list[_ParsedRow] = []
    seen_empl_no: dict[str, int] = {}

    for offset, raw_row in enumerate(data_rows):
        row_no = offset + 2  # 1행은 헤더, 데이터는 2행부터
        if raw_row is None or all(v is None for v in raw_row):
            continue  # 완전히 빈 행은 건너뜀 (Excel 저장 시 흔히 생기는 trailing 빈 행)

        empl_no = _cell(raw_row, 0)
        empl_nm = _cell(raw_row, 1)
        dept_nm = _cell(raw_row, 2)
        jikgup_nm = _cell(raw_row, 3)
        roles_raw = _cell(raw_row, 4)
        skills_raw = _cell(raw_row, 5)
        prfcy_raw = raw_row[6] if len(raw_row) > 6 else None
        hire_dt_raw = raw_row[7] if len(raw_row) > 7 else None
        empl_stat_label = _cell(raw_row, 8)
        mphone_no = _cell(raw_row, 9)

        row_error_count_before = len(errors)

        if not empl_no:
            errors.append(_err(row_no, "사번", empl_no, "필수 항목입니다."))
        elif empl_no in seen_empl_no:
            errors.append(_err(row_no, "사번", empl_no, f"파일 내 {seen_empl_no[empl_no]}행과 사번이 중복됩니다."))
        elif empl_no not in existing_empl_nos:
            errors.append(
                _err(
                    row_no,
                    "사번",
                    empl_no,
                    "신규 사원은 이메일이 필수(2026-07-06 정책)라 Excel Import로 등록할 수 없습니다 — "
                    "사원 목록 화면에서 개별 등록하세요.",
                )
            )
        else:
            seen_empl_no[empl_no] = row_no

        if not empl_nm:
            errors.append(_err(row_no, "성명", empl_nm, "필수 항목입니다."))

        dept_id = None
        if not dept_nm:
            errors.append(_err(row_no, "팀", dept_nm, "필수 항목입니다."))
        elif dept_nm not in dept_by_name:
            errors.append(_err(row_no, "팀", dept_nm, "등록되지 않은 부서명입니다."))
        else:
            dept_id = dept_by_name[dept_nm]

        jikgup_id = None
        if not jikgup_nm:
            errors.append(_err(row_no, "직급", jikgup_nm, "필수 항목입니다."))
        elif jikgup_nm not in jikgup_by_name:
            errors.append(_err(row_no, "직급", jikgup_nm, "등록되지 않은 직급명입니다."))
        else:
            jikgup_id = jikgup_by_name[jikgup_nm]

        jikmu_ids: list[uuid.UUID] = []
        if roles_raw:
            for code in (c.strip() for c in roles_raw.split(",")):
                if not code:
                    continue
                if code not in jikmu_by_code:
                    errors.append(_err(row_no, "보유역할", code, "등록되지 않은 역할 코드(JIKMU_CD)입니다."))
                else:
                    jikmu_ids.append(jikmu_by_code[code])

        skill_ids: list[uuid.UUID] = []
        if skills_raw:
            for name in (s.strip() for s in skills_raw.split(",")):
                if not name:
                    continue
                if name not in skill_by_name:
                    errors.append(_err(row_no, "주요기술", name, "등록되지 않은 기술명입니다."))
                else:
                    skill_ids.append(skill_by_name[name])

        prfcy_levl = None
        if prfcy_raw is not None and str(prfcy_raw).strip() != "":
            try:
                prfcy_levl = int(prfcy_raw)
            except (TypeError, ValueError):
                errors.append(_err(row_no, "숙련도", prfcy_raw, "숫자여야 합니다."))
            else:
                if not 1 <= prfcy_levl <= 5:
                    errors.append(_err(row_no, "숙련도", prfcy_raw, "1~5 범위여야 합니다."))

        hire_dt = None
        if hire_dt_raw is not None and str(hire_dt_raw).strip() != "":
            if isinstance(hire_dt_raw, datetime):
                hire_dt = hire_dt_raw.date()
            elif isinstance(hire_dt_raw, date):
                hire_dt = hire_dt_raw
            else:
                try:
                    hire_dt = date.fromisoformat(str(hire_dt_raw).strip())
                except ValueError:
                    errors.append(_err(row_no, "입사일", hire_dt_raw, "YYYY-MM-DD 형식이어야 합니다."))

        empl_stat_cd = "ACTIVE"
        if not empl_stat_label:
            errors.append(_err(row_no, "재직상태", empl_stat_label, "필수 항목입니다."))
        elif empl_stat_label not in _EMPL_STAT_LABEL_TO_CD:
            errors.append(_err(row_no, "재직상태", empl_stat_label, "재직/휴직/퇴직 중 하나여야 합니다."))
        else:
            empl_stat_cd = _EMPL_STAT_LABEL_TO_CD[empl_stat_label]

        if len(errors) == row_error_count_before:
            parsed_rows.append(
                _ParsedRow(
                    row_no=row_no,
                    empl_no=empl_no,
                    empl_nm=empl_nm,
                    dept_id=dept_id,
                    jikgup_id=jikgup_id,
                    jikmu_ids=jikmu_ids,
                    skill_ids=skill_ids,
                    prfcy_levl=prfcy_levl,
                    hire_dt=hire_dt,
                    empl_stat_cd=empl_stat_cd,
                    mphone_no=mphone_no,
                )
            )

    total_rows = len(data_rows)
    if errors:
        raise EmployeeImportValidationError(total_rows, errors)
    return parsed_rows


def _fail_header(all_rows: list[tuple], header: list[str]) -> None:
    raise EmployeeImportValidationError(
        max(len(all_rows) - 1, 0),
        [_err(1, "header", ", ".join(header), f"헤더가 예상 형식과 다릅니다. 필요한 헤더: {', '.join(_HEADERS)}")],
    )


def apply_import(db: Session, rows: list[_ParsedRow]) -> dict:
    """검증을 통과한 행들을 하나의 트랜잭션으로 Upsert한다. 보유역할·기술스택은
    사원별로 Import 파일 기준으로 전체 동기화(파일에 없는 기존 관계는 삭제)한다."""
    created = 0
    updated = 0
    roles_processed = 0
    skills_processed = 0

    try:
        for row in rows:
            employee = db.scalar(select(HrEmplMst).where(HrEmplMst.EMPL_NO == row.empl_no))
            if employee is None:
                employee = HrEmplMst(EMPL_NO=row.empl_no)
                db.add(employee)
                created += 1
            else:
                updated += 1

            employee.EMPL_NM = row.empl_nm
            employee.DEPT_ID = row.dept_id
            employee.JIKGUP_ID = row.jikgup_id
            employee.HIRE_DT = row.hire_dt
            employee.EMPL_STAT_CD = row.empl_stat_cd
            employee.MPHONE_NO = row.mphone_no
            db.flush()  # 신규 사원의 EMPL_ID를 아래 역할/기술 연결에 사용하기 위해 필요

            db.execute(delete(HrEmplRoleRel).where(HrEmplRoleRel.EMPL_ID == employee.EMPL_ID))
            for jikmu_id in row.jikmu_ids:
                db.add(HrEmplRoleRel(EMPL_ID=employee.EMPL_ID, JIKMU_ID=jikmu_id))
                roles_processed += 1

            db.execute(delete(HrEmplSkillRel).where(HrEmplSkillRel.EMPL_ID == employee.EMPL_ID))
            for skill_id in row.skill_ids:
                db.add(HrEmplSkillRel(EMPL_ID=employee.EMPL_ID, SKILL_ID=skill_id, PRFCY_LEVL=row.prfcy_levl))
                skills_processed += 1

        db.commit()
    except Exception:
        db.rollback()
        raise

    return {
        "total_rows": len(rows),
        "created_count": created,
        "updated_count": updated,
        "roles_processed_count": roles_processed,
        "skills_processed_count": skills_processed,
    }
