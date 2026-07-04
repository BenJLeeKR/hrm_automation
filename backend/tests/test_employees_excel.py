import io

import openpyxl

# GET /export가 GET /{empl_id}(UUID 경로)보다 뒤에 등록되어 "export"가 UUID로 파싱
# 시도되어 422로 가로채이던 라우팅 순서 버그(2026-07-03 Excel Import/Export UI 작업 중
# 발견 및 수정)의 회귀를 방지하기 위한 테스트.


def test_export_returns_xlsx(client, admin_token, dept, jikgup):
    resp = client.get("/api/v1/employees/export", headers={"Authorization": f"Bearer {admin_token}"})
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert headers == ["사번", "성명", "팀", "직급", "보유역할", "주요기술", "숙련도", "입사일", "재직상태", "휴대폰번호"]


def test_import_invalid_file_returns_422(client, admin_token):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["사번", "성명", "팀", "직급", "보유역할", "주요기술", "숙련도", "입사일", "재직상태", "휴대폰번호"])
    ws.append(["", "홍길동", "없는팀", "없는직급", "", "", "", "", "", ""])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = client.post(
        "/api/v1/employees/import",
        headers={"Authorization": f"Bearer {admin_token}"},
        files={"file": ("test.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 422, resp.text
    body = resp.json()["detail"]
    assert body["error_count"] > 0
    assert body["errors"][0]["row"] == 2
